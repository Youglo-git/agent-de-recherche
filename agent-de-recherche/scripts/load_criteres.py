"""Charge et valide criteres.xlsx en un objet Criteres typé.

Le but est de *casser tôt* si le fichier est incohérent, plutôt que de laisser
une recherche tourner pendant dix minutes pour produire des résultats vides.
Toute erreur de structure remonte explicitement avec le numéro de ligne.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook


Poids = Literal["obligatoire", "souhaitable", "exclu"]
PoidsProfil = Literal["obligatoire", "souhaitable", "informatif"]
Operateur = Literal[">=", "<=", "=", "in", "not_in"]
Methode = Literal["API", "Chrome"]


@dataclass
class MotCle:
    mot_cle: str
    categorie: str
    poids: Poids
    commentaire: str = ""


@dataclass
class Localisation:
    ville: str
    region: str
    rayon_km: int
    teletravail_min_pct: int
    actif: bool


@dataclass
class CritereProfil:
    """Un critère structuré du profil ou du poste.

    `valeur` est conservée sous sa forme originale (str/int/float) ; les
    consommateurs (matcher, scoring) interprètent selon `operateur`.
    Pour `in` / `not_in`, la valeur est une liste de chaînes (issue d'un split
    sur la virgule).
    """
    critere: str
    valeur: object
    operateur: Operateur
    poids: PoidsProfil
    commentaire: str = ""


@dataclass
class Source:
    source: str
    methode: Methode
    actif: bool
    notes: str = ""


@dataclass
class Criteres:
    mots_cles: list[MotCle]
    localisations: list[Localisation]
    profil_poste: list[CritereProfil]
    sources: list[Source]

    def obligatoires(self) -> list[MotCle]:
        return [m for m in self.mots_cles if m.poids == "obligatoire"]

    def souhaitables(self) -> list[MotCle]:
        return [m for m in self.mots_cles if m.poids == "souhaitable"]

    def exclusions(self) -> list[MotCle]:
        return [m for m in self.mots_cles if m.poids == "exclu"]

    def profil_obligatoires(self) -> list[CritereProfil]:
        return [c for c in self.profil_poste if c.poids == "obligatoire"]

    def profil_souhaitables(self) -> list[CritereProfil]:
        return [c for c in self.profil_poste if c.poids == "souhaitable"]

    def profil_informatifs(self) -> list[CritereProfil]:
        return [c for c in self.profil_poste if c.poids == "informatif"]

    def critere_par_nom(self, nom: str) -> CritereProfil | None:
        """Accès direct par nom (ex: 'salaire_annuel_brut_min_eur')."""
        for c in self.profil_poste:
            if c.critere == nom:
                return c
        return None

    def sources_actives(self) -> list[Source]:
        return [s for s in self.sources if s.actif]


class CriteresError(ValueError):
    """Erreur de chargement du fichier de critères."""


def _require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise CriteresError(
            f"Feuille manquante : '{name}'. Régénère le template avec init_criteres.py."
        )
    return wb[name]


def _rows(ws, skip_header: bool = True):
    """Itère sur les lignes non vides d'une feuille."""
    it = ws.iter_rows(values_only=True)
    if skip_header:
        next(it, None)
    for row_idx, row in enumerate(it, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue
        yield row_idx, row


def _to_bool(value, row_idx: int, col: str) -> bool:
    """Convertit OUI/NON (avec tolérance accents/casse) en booléen."""
    if value is None:
        raise CriteresError(f"Ligne {row_idx} colonne {col}: valeur vide attendue OUI/NON.")
    v = str(value).strip().upper()
    if v in {"OUI", "YES", "TRUE", "1"}:
        return True
    if v in {"NON", "NO", "FALSE", "0"}:
        return False
    raise CriteresError(f"Ligne {row_idx} colonne {col}: '{value}' n'est ni OUI ni NON.")


def _load_mots_cles(ws) -> list[MotCle]:
    out: list[MotCle] = []
    valid_poids = {"obligatoire", "souhaitable", "exclu"}
    for row_idx, (mc, cat, poids, com) in _rows(ws):
        if not mc:
            continue
        poids_norm = str(poids or "").strip().lower()
        if poids_norm not in valid_poids:
            raise CriteresError(
                f"Mots_cles ligne {row_idx}: poids invalide '{poids}'. "
                f"Attendu: {sorted(valid_poids)}."
            )
        out.append(
            MotCle(
                mot_cle=str(mc).strip(),
                categorie=str(cat or "").strip(),
                poids=poids_norm,  # type: ignore[arg-type]
                commentaire=str(com or "").strip(),
            )
        )
    if not out:
        raise CriteresError("Aucun mot-clé renseigné dans la feuille Mots_cles.")
    return out


def _load_localisations(ws) -> list[Localisation]:
    out: list[Localisation] = []
    for row_idx, row in _rows(ws):
        ville, region, rayon, teletravail, actif = (list(row) + [None] * 5)[:5]
        if not ville and not str(region or "").strip():
            continue
        try:
            rayon_i = int(rayon) if rayon not in (None, "") else 0
            tt_i = int(teletravail) if teletravail not in (None, "") else 0
        except ValueError as err:
            raise CriteresError(
                f"Localisation ligne {row_idx}: rayon_km et teletravail_min_pct "
                f"doivent être des entiers ({err})."
            ) from err
        out.append(
            Localisation(
                ville=str(ville or "").strip(),
                region=str(region or "").strip(),
                rayon_km=rayon_i,
                teletravail_min_pct=tt_i,
                actif=_to_bool(actif, row_idx, "actif"),
            )
        )
    return out


_OPERATEURS_VALIDES = {">=", "<=", "=", "in", "not_in"}
_POIDS_PROFIL_VALIDES = {"obligatoire", "souhaitable", "informatif"}


def _coerce_valeur(raw, operateur: str, critere: str, row_idx: int) -> object:
    """Convertit la valeur brute Excel selon l'opérateur attendu.

    - >= et <=  attendent un nombre (int ou float)
    - =         accepte nombre ou texte
    - in, not_in attendent une liste (issue d'un split sur la virgule)
    """
    if raw is None or raw == "":
        # On tolère vide pour 'informatif' ; lève pour les autres poids au niveau appelant.
        if operateur in {"in", "not_in"}:
            return []
        return None

    if operateur in {">=", "<="}:
        try:
            # Conserver int si entier exact, sinon float
            if isinstance(raw, (int, float)):
                return raw
            return float(str(raw).replace(",", ".").strip())
        except (TypeError, ValueError) as err:
            raise CriteresError(
                f"Profil_Poste ligne {row_idx}: critère '{critere}' avec opérateur "
                f"'{operateur}' attend un nombre, reçu '{raw}'."
            ) from err

    if operateur in {"in", "not_in"}:
        if isinstance(raw, (int, float)):
            return [str(raw).strip()]
        items = [x.strip() for x in str(raw).split(",") if x.strip()]
        return items

    # operateur "="
    return raw if isinstance(raw, (int, float)) else str(raw).strip()


def _load_profil_poste(ws) -> list[CritereProfil]:
    """Charge la feuille Profil_Poste (structure unifiée)."""
    out: list[CritereProfil] = []
    deja_vus: set[str] = set()

    for row_idx, row in _rows(ws):
        critere, valeur, operateur, poids, commentaire = (list(row) + [None] * 5)[:5]
        if not critere:
            continue

        critere = str(critere).strip()
        if critere in deja_vus:
            raise CriteresError(
                f"Profil_Poste ligne {row_idx}: critère '{critere}' déjà défini. "
                f"Supprime le doublon pour éviter toute ambiguïté."
            )
        deja_vus.add(critere)

        op_norm = str(operateur or "").strip()
        if op_norm not in _OPERATEURS_VALIDES:
            raise CriteresError(
                f"Profil_Poste ligne {row_idx}: opérateur invalide '{operateur}'. "
                f"Attendu: {sorted(_OPERATEURS_VALIDES)}."
            )

        poids_norm = str(poids or "").strip().lower()
        if poids_norm not in _POIDS_PROFIL_VALIDES:
            raise CriteresError(
                f"Profil_Poste ligne {row_idx}: poids invalide '{poids}'. "
                f"Attendu: {sorted(_POIDS_PROFIL_VALIDES)}."
            )

        valeur_coerce = _coerce_valeur(valeur, op_norm, critere, row_idx)

        # 'obligatoire' ne peut pas avoir de valeur vide — casser tôt
        if poids_norm == "obligatoire" and (
            valeur_coerce is None or (isinstance(valeur_coerce, list) and not valeur_coerce)
        ):
            raise CriteresError(
                f"Profil_Poste ligne {row_idx}: critère obligatoire '{critere}' "
                f"a une valeur vide. Renseigne une valeur ou passe le poids à 'informatif'."
            )

        out.append(
            CritereProfil(
                critere=critere,
                valeur=valeur_coerce,
                operateur=op_norm,  # type: ignore[arg-type]
                poids=poids_norm,  # type: ignore[arg-type]
                commentaire=str(commentaire or "").strip(),
            )
        )

    if not out:
        raise CriteresError("Aucun critère renseigné dans la feuille Profil_Poste.")
    return out


def _load_sources(ws) -> list[Source]:
    out: list[Source] = []
    valid_methodes = {"API", "Chrome"}
    for row_idx, row in _rows(ws):
        source, methode, actif, notes = (list(row) + [None] * 4)[:4]
        if not source:
            continue
        methode_norm = str(methode or "").strip()
        if methode_norm not in valid_methodes:
            raise CriteresError(
                f"Sources ligne {row_idx}: methode invalide '{methode}'. "
                f"Attendu: {sorted(valid_methodes)}."
            )
        out.append(
            Source(
                source=str(source).strip(),
                methode=methode_norm,  # type: ignore[arg-type]
                actif=_to_bool(actif, row_idx, "actif"),
                notes=str(notes or "").strip(),
            )
        )
    if not out:
        raise CriteresError("Aucune source renseignée dans la feuille Sources.")
    return out


def load_criteres(path: Path) -> Criteres:
    """Charge et valide le fichier Excel des critères.

    Lève CriteresError en cas d'incohérence — avec un message pointant la ligne.
    """
    if not path.exists():
        raise CriteresError(
            f"Fichier introuvable : {path}. Lance d'abord init_criteres.py."
        )

    wb = load_workbook(path, data_only=True, read_only=True)

    try:
        criteres = Criteres(
            mots_cles=_load_mots_cles(_require_sheet(wb, "Mots_cles")),
            localisations=_load_localisations(_require_sheet(wb, "Localisation")),
            profil_poste=_load_profil_poste(_require_sheet(wb, "Profil_Poste")),
            sources=_load_sources(_require_sheet(wb, "Sources")),
        )
    finally:
        wb.close()

    return criteres


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", type=Path, default=Path("criteres.xlsx"))
    parser.add_argument(
        "--json",
        action="store_true",
        help="Afficher le résultat en JSON (utile pour debug / autres scripts).",
    )
    args = parser.parse_args()

    try:
        criteres = load_criteres(args.path)
    except CriteresError as err:
        print(f"[ERREUR] {err}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(asdict(criteres), ensure_ascii=False, indent=2))
        return 0

    print(f"[OK] Critères chargés depuis {args.path}")
    print(f"     Mots-clés         : {len(criteres.mots_cles)}"
          f"   (obligatoires: {len(criteres.obligatoires())},"
          f" souhaitables: {len(criteres.souhaitables())},"
          f" exclus: {len(criteres.exclusions())})")
    print(f"     Localisations     : {len(criteres.localisations)}")
    print(f"     Profil_Poste      : {len(criteres.profil_poste)}"
          f"   (obligatoires: {len(criteres.profil_obligatoires())},"
          f" souhaitables: {len(criteres.profil_souhaitables())},"
          f" informatifs: {len(criteres.profil_informatifs())})")
    print(f"     Sources actives   : {len(criteres.sources_actives())}/{len(criteres.sources)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
