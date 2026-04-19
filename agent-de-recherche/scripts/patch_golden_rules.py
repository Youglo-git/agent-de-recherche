"""Migration idempotente du `criteres.xlsx` existant vers les Golden Rules.

Ce patch ne recrée PAS le fichier (ce qui écraserait les personnalisations) :
il modifie **uniquement** les cellules qui doivent évoluer suite à l'introduction
des Golden Rules :

- Feuille `Profil_Poste`, ligne `salaire_annuel_brut_min_eur` :
  - colonne `poids`       : `obligatoire` → `souhaitable`
  - colonne `commentaire` : mis à jour pour expliquer la Golden Rule #3

Idempotent : si la ligne est déjà à `souhaitable`, on ne touche à rien.
Les autres feuilles et lignes ne sont pas modifiées — les validations de données
(listes déroulantes Excel) sont préservées par openpyxl.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


NOUVEAU_COMMENTAIRE = (
    "Golden Rule #3 : NON bloquant. Si l'annonce ne mentionne aucun "
    "salaire, elle est conservée. N'écarte que les annonces affichant "
    "explicitement un salaire inférieur au plancher."
)


def patcher(path: Path) -> dict[str, object]:
    """Applique le patch et retourne un résumé {'modifie': bool, 'raisons': [...]}."""
    if not path.exists():
        raise FileNotFoundError(f"{path} introuvable")

    wb = load_workbook(path)
    if "Profil_Poste" not in wb.sheetnames:
        raise ValueError("Feuille 'Profil_Poste' introuvable — fichier non standard.")

    ws = wb["Profil_Poste"]
    modifs: list[str] = []

    # Colonnes attendues : A=critere B=valeur C=operateur D=poids E=commentaire
    for row in ws.iter_rows(min_row=2, max_col=5):
        critere_cell = row[0]
        if (critere_cell.value or "").strip().lower() != "salaire_annuel_brut_min_eur":
            continue

        poids_cell = row[3]
        commentaire_cell = row[4]

        if (poids_cell.value or "").strip().lower() == "obligatoire":
            poids_cell.value = "souhaitable"
            modifs.append(f"ligne {critere_cell.row} : poids obligatoire → souhaitable")

        if commentaire_cell.value != NOUVEAU_COMMENTAIRE:
            commentaire_cell.value = NOUVEAU_COMMENTAIRE
            modifs.append(f"ligne {critere_cell.row} : commentaire mis à jour")

        break
    else:
        modifs.append("Critère 'salaire_annuel_brut_min_eur' introuvable — rien à faire.")

    if modifs and not all(m.startswith("Critère ") for m in modifs):
        wb.save(path)
        return {"modifie": True, "raisons": modifs}
    return {"modifie": False, "raisons": modifs}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--path",
        type=Path,
        required=True,
        help="Chemin du fichier criteres.xlsx à migrer.",
    )
    args = parser.parse_args()

    try:
        res = patcher(args.path)
    except (FileNotFoundError, ValueError) as err:
        print(f"[ERREUR] {err}", file=sys.stderr)
        return 1

    if res["modifie"]:
        print(f"[OK] {args.path} migré vers Golden Rules :")
        for r in res["raisons"]:
            print(f"  - {r}")
    else:
        print(f"[OK] {args.path} déjà conforme aux Golden Rules — aucune modification.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
