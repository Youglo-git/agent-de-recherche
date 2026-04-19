"""Initialise le fichier criteres.xlsx avec un template pré-rempli.

Le fichier contient 4 feuilles :
- Mots_cles      : mots-clés métier/ATS avec poids (obligatoire/souhaitable/exclu)
- Localisation   : villes acceptées, rayon, % télétravail min (détail des pôles)
- Profil_Poste   : critères structurés du profil et du poste — structure unifiée
                   (critere, valeur, operateur, poids, commentaire)
- Sources        : activation/désactivation par site

Idempotent : refuse d'écraser un fichier existant.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, columns: list[str]) -> None:
    """Applique un style homogène à la ligne d'en-tête."""
    for idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(idx)].width = max(18, len(title) + 4)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _build_mots_cles(ws) -> None:
    """Feuille des mots-clés métier et ATS avec pondération."""
    columns = ["mot_cle", "categorie", "poids", "commentaire"]
    _style_header(ws, columns)

    exemples = [
        ("Directeur de projet", "Intitule", "obligatoire", "Titre de poste cible"),
        ("Chef de programme", "Intitule", "souhaitable", "Variante équivalente"),
        ("Program Manager", "Intitule", "souhaitable", "Variante anglophone"),
        ("Prince2", "Methodologie", "souhaitable", "Certification détenue"),
        ("Scrum", "Methodologie", "souhaitable", "Certification PO"),
        ("Agile", "Methodologie", "souhaitable", ""),
        ("PMO", "Fonction", "souhaitable", ""),
        ("Transformation", "Domaine", "souhaitable", ""),
        ("SI", "Domaine", "souhaitable", "Systèmes d'information"),
        ("Cloud", "Techno", "souhaitable", ""),
        ("Stagiaire", "Autre", "exclu", "Ne pas retenir les annonces stagiaire"),
        ("Alternant", "Autre", "exclu", ""),
    ]
    for row_idx, (mc, cat, poids, com) in enumerate(exemples, start=2):
        ws.cell(row=row_idx, column=1, value=mc)
        ws.cell(row=row_idx, column=2, value=cat)
        ws.cell(row=row_idx, column=3, value=poids)
        ws.cell(row=row_idx, column=4, value=com)

    # Liste déroulante sur poids
    dv = DataValidation(
        type="list",
        formula1='"obligatoire,souhaitable,exclu"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Valeurs autorisées : obligatoire, souhaitable, exclu"
    dv.errorTitle = "Poids invalide"
    ws.add_data_validation(dv)
    dv.add(f"C2:C1000")


def _build_localisation(ws) -> None:
    """Feuille localisation et télétravail."""
    columns = ["ville", "region", "rayon_km", "teletravail_min_pct", "actif"]
    _style_header(ws, columns)

    exemples = [
        ("Paris", "Île-de-France", 30, 60, "OUI"),
        ("Lyon", "Auvergne-Rhône-Alpes", 20, 80, "OUI"),
        ("Full remote", "", 0, 100, "OUI"),
    ]
    for row_idx, row in enumerate(exemples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    dv = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("E2:E1000")


def _build_profil_poste(ws) -> None:
    """Feuille des critères structurés profil + poste.

    Structure unifiée pour tous les critères chiffrés / catégoriels :
    - critere     : identifiant technique (snake_case)
    - valeur      : valeur attendue (nombre, texte, ou liste séparée par ',')
    - operateur   : >=, <=, =, in, not_in
    - poids       : obligatoire | souhaitable | informatif
    - commentaire : note libre pour l'utilisateur

    'informatif' = pas pris en compte dans le filtre, mais conservé pour le
    scoring et le rapport (skill suivante du pipeline).
    """
    columns = ["critere", "valeur", "operateur", "poids", "commentaire"]
    _style_header(ws, columns)

    # Format : (critere, valeur, operateur, poids, commentaire)
    exemples = [
        # — Rémunération & contrat —
        ("salaire_annuel_brut_min_eur", 85000, ">=", "souhaitable",
         "Golden Rule #3 : NON bloquant. Si l'annonce ne mentionne aucun "
         "salaire, elle est conservée. N'écarte que les annonces affichant "
         "explicitement un salaire inférieur au plancher."),
        ("salaire_annuel_brut_cible_eur", 110000, ">=", "informatif",
         "Cible marché senior, indicatif"),
        ("contrat_accepte", "CDI,Freelance", "in", "obligatoire",
         "Types de contrat acceptés (séparés par virgule)"),

        # — Expérience —
        ("experience_min_annees", 10, ">=", "obligatoire",
         "Expérience minimale demandée par l'annonce"),
        ("experience_max_annees", 25, "<=", "souhaitable",
         "Annonces ciblant 'junior 0-3 ans' à exclure"),

        # — Profil cadre —
        ("statut", "Cadre,Cadre dirigeant", "in", "obligatoire",
         "Écarte les postes 'agent de maîtrise' et techniciens"),
        ("anglais_niveau_accepte", "B2,C1,C2,Bilingue", "in", "souhaitable",
         "Niveaux CECRL acceptés (équivalent : B2 ou plus). Mettre A1,A2,B1 si ouvert à débutants"),
        ("autres_langues", "", "in", "informatif",
         "Optionnel, ex: 'Allemand,Espagnol'"),

        # — Mobilité & télétravail —
        ("teletravail_min_pct", 60, ">=", "souhaitable",
         "Pourcentage min de télétravail accepté"),
        ("deplacements_max_pct", 30, "<=", "souhaitable",
         "Pourcentage max de déplacements accepté"),
        ("mobilite_internationale", "Non", "=", "souhaitable",
         "Oui / Non — refus des postes nécessitant expatriation"),
        ("zone_geo_acceptee", "France,Belgique,Suisse,Luxembourg", "in", "souhaitable",
         "Pays acceptés pour la localisation du poste"),

        # — Environnement entreprise —
        ("taille_entreprise_acceptee", "ETI,Grand Groupe,Scale-up", "in", "souhaitable",
         "Catégories : TPE, PME, ETI, Grand Groupe, Start-up, Scale-up"),
        ("taille_equipe_a_manager_min", 0, ">=", "informatif",
         "Nombre de personnes à manager (direct + indirect), 0 = pas de critère"),
        ("budget_a_gerer_min_meur", 0, ">=", "informatif",
         "Budget annuel à piloter en M€, 0 = pas de critère"),
    ]
    for row_idx, row in enumerate(exemples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Listes déroulantes pour opérateur et poids
    dv_op = DataValidation(type="list", formula1='">=,<=,=,in,not_in"', allow_blank=False)
    dv_op.error = "Opérateurs valides : >=, <=, =, in, not_in"
    dv_op.errorTitle = "Opérateur invalide"
    ws.add_data_validation(dv_op)
    dv_op.add("C2:C1000")

    dv_poids = DataValidation(
        type="list",
        formula1='"obligatoire,souhaitable,informatif"',
        allow_blank=False,
    )
    dv_poids.error = "Poids valides : obligatoire, souhaitable, informatif"
    dv_poids.errorTitle = "Poids invalide"
    ws.add_data_validation(dv_poids)
    dv_poids.add("D2:D1000")

    # Élargir les colonnes pour la lisibilité
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 50


def _build_sources(ws) -> None:
    """Feuille d'activation par source."""
    columns = ["source", "methode", "actif", "notes"]
    _style_header(ws, columns)

    exemples = [
        ("France Travail", "API", "OUI", "API officielle, nécessite credentials OAuth"),
        ("Adzuna", "API", "OUI", "API REST gratuite, agrège plusieurs sources"),
        ("LinkedIn", "Chrome", "OUI", "Scraping via navigateur connecté — risque CGU"),
        ("APEC", "Chrome", "OUI", "Idéal cadres"),
        ("Welcome to the Jungle", "Chrome", "OUI", "Tech / scale-ups"),
        ("HelloWork", "Chrome", "OUI", "Agrégateur généraliste"),
        ("Indeed", "Chrome", "NON", "Déjà couvert par Adzuna — activer uniquement si manque"),
    ]
    for row_idx, row in enumerate(exemples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    dv = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("C2:C1000")


def create_template(path: Path, *, force: bool = False) -> None:
    """Crée le fichier Excel de paramétrage.

    Args:
        path: chemin de destination du .xlsx
        force: si False (défaut), refuse d'écraser un fichier existant.

    Raises:
        FileExistsError: si le fichier existe déjà et force=False.
    """
    if path.exists() and not force:
        raise FileExistsError(
            f"{path} existe déjà. Supprime-le manuellement ou relance avec --force "
            f"(attention : cela écrase tes paramètres actuels)."
        )

    wb = Workbook()
    # openpyxl crée une feuille par défaut — on la renomme
    ws = wb.active
    ws.title = "Mots_cles"
    _build_mots_cles(ws)

    _build_localisation(wb.create_sheet("Localisation"))
    _build_profil_poste(wb.create_sheet("Profil_Poste"))
    _build_sources(wb.create_sheet("Sources"))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--path",
        type=Path,
        default=Path("criteres.xlsx"),
        help="Chemin du fichier à créer (défaut: ./criteres.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Écraser le fichier s'il existe déjà (⚠ perte des paramètres).",
    )
    args = parser.parse_args()

    try:
        create_template(args.path, force=args.force)
    except FileExistsError as err:
        print(f"[ERREUR] {err}", file=sys.stderr)
        return 1

    print(f"[OK] Template créé : {args.path.resolve()}")
    print("     Ouvre-le dans Excel pour adapter les critères à ta recherche.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
